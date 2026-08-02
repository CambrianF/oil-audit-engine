import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from src.auditor import audit_invoice


def save_workbook_safely(wb, path):
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        fallback = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        wb.save(fallback)
        print(f"WARNING: {path} was locked (likely open in Excel). Saved to {fallback} instead.")
        return fallback


def main():
    invoice_dir = "data/sample_invoices"
    print(f"Starting Executive Oil Audit Engine on directory: {invoice_dir}\n")

    if not os.path.exists(invoice_dir):
        print(f"Error: Directory '{invoice_dir}' does not exist.")
        return

    files = [os.path.join(invoice_dir, f) for f in os.listdir(invoice_dir) if f.endswith(".pdf")]

    if not files:
        print("No PDF invoices found to audit.")
        return

    audit_results = []

    for file_path in files:
        result = audit_invoice(file_path)
        audit_results.append({
            "File Name": result["file"],
            "Vendor Name": result.get("vendor_name", "Unknown Vendor"),
            "Account #": result.get("account_number", "N/A"),
            "Invoice #": result.get("invoice_number", "N/A"),
            "AFE #": result.get("afe_number", "N/A"),
            "API #": result.get("api_number", "N/A"),
            "Audit Status": result["status"],
            "Financial Issues": ", ".join(result["financial_issues"]),
            "Compliance Issues": ", ".join(result["compliance_issues"]),
            "Vendor Risk Level": result.get("vendor_risk_level", "N/A"),
            "Snippet Preview": result["raw_text_snippet"]
        })
        print("----------------------------------------")
        print(f"File: {result['file']}")
        print(f"Vendor: {result.get('vendor_name', 'Unknown Vendor')}")
        print(f"Account #: {result.get('account_number', 'N/A')} | Invoice #: {result.get('invoice_number', 'N/A')}")
        print(f"AFE #: {result.get('afe_number', 'N/A')} | API #: {result.get('api_number', 'N/A')}")
        print(f"Status: {result['status']}")
        print(f"Financial Issues: {result['financial_issues']}")
        print(f"Compliance Issues: {result['compliance_issues']}")
        print(f"Vendor Risk Level: {result.get('vendor_risk_level', 'N/A')}")
        print("----------------------------------------\n")

    total_invoices = len(audit_results)
    passed_invoices = sum(1 for r in audit_results if r["Audit Status"] == "Passed")
    review_invoices = sum(1 for r in audit_results if r["Audit Status"] == "Review")
    flagged_invoices = sum(1 for r in audit_results if r["Audit Status"] == "Flagged")
    pass_rate = (passed_invoices / total_invoices) if total_invoices > 0 else 0.0

    excel_output = "oil_audit_summary.xlsx"

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    title_font = Font(name="Calibri", size=16, bold=True, color="1B2A4A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="555555")
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10, color="000000")
    bold_cell_font = Font(name="Calibri", size=10, bold=True, color="000000")

    thin_border = Border(
        left=Side(border_style="thin", color="D3D3D3"),
        right=Side(border_style="thin", color="D3D3D3"),
        top=Side(border_style="thin", color="D3D3D3"),
        bottom=Side(border_style="thin", color="D3D3D3")
    )

    ws_summary["A1"] = "Oilfield Vendor Audit - Executive Dashboard"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = "Portfolio Health & Compliance Telemetry"
    ws_summary["A2"].font = subtitle_font

    ws_summary["A4"] = "Portfolio Metric"
    ws_summary["B4"] = "Value"
    ws_summary.row_dimensions[4].height = 26
    for col_letter in ["A", "B"]:
        cell = ws_summary[f"{col_letter}4"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left" if col_letter == "A" else "center", vertical="center")
        cell.border = thin_border

    kpi_data = [
        ("Total Invoices Audited", total_invoices),
        ("Passed Compliance", passed_invoices),
        ("Review (Informational)", review_invoices),
        ("Flagged / Discrepancies", flagged_invoices),
        ("Overall Compliance Pass Rate", f"{pass_rate:.1%}")
    ]

    for idx, (metric, val) in enumerate(kpi_data, 5):
        ws_summary.row_dimensions[idx].height = 22
        c1 = ws_summary.cell(row=idx, column=1, value=metric)
        c1.font = bold_cell_font
        c1.border = thin_border
        c1.alignment = Alignment(horizontal="left", vertical="center")

        c2 = ws_summary.cell(row=idx, column=2, value=val)
        c2.font = cell_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="center", vertical="center")

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    ws_log = wb.create_sheet(title="Detailed Audit Log")
    ws_log.views.sheetView[0].showGridLines = True

    ws_log["A1"] = "Itemized Invoice Audit Log"
    ws_log["A1"].font = title_font
    ws_log["A2"] = "Complete audit trail of processed PDF vendor documents"
    ws_log["A2"].font = subtitle_font

    headers = ["File Name", "Vendor Name", "Account #", "Invoice #", "AFE #", "API #",
               "Audit Status", "Financial Issues", "Compliance Issues", "Vendor Risk Level", "Snippet Preview"]
    ws_log.row_dimensions[4].height = 26
    for col_idx, header in enumerate(headers, 1):
        cell = ws_log.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_idx == 7 else "left", vertical="center", wrap_text=True)
        cell.border = thin_border

    pass_font = Font(name="Calibri", size=10, bold=True, color="006100")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    review_font = Font(name="Calibri", size=10, bold=True, color="9C6500")
    review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    flag_font = Font(name="Calibri", size=10, bold=True, color="9C0006")
    flag_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    risk_high_font = Font(name="Calibri", size=9, bold=True, color="9C0006")
    risk_moderate_font = Font(name="Calibri", size=9, bold=True, color="9C6500")
    risk_low_font = Font(name="Calibri", size=9, color="006100")
    risk_default_font = Font(name="Calibri", size=9, italic=True, color="555555")

    for row_idx, row_data in enumerate(audit_results, 5):
        ws_log.row_dimensions[row_idx].height = 45

        c1 = ws_log.cell(row=row_idx, column=1, value=row_data["File Name"])
        c1.font = cell_font
        c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c1.border = thin_border

        c2 = ws_log.cell(row=row_idx, column=2, value=row_data["Vendor Name"])
        c2.font = bold_cell_font
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c2.border = thin_border

        c3 = ws_log.cell(row=row_idx, column=3, value=row_data["Account #"])
        c3.font = cell_font
        c3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c3.border = thin_border

        c4 = ws_log.cell(row=row_idx, column=4, value=row_data["Invoice #"])
        c4.font = cell_font
        c4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c4.border = thin_border

        c5 = ws_log.cell(row=row_idx, column=5, value=row_data["AFE #"])
        c5.font = cell_font
        c5.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c5.border = thin_border

        c6 = ws_log.cell(row=row_idx, column=6, value=row_data["API #"])
        c6.font = cell_font
        c6.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c6.border = thin_border

        status_val = row_data["Audit Status"]
        c7 = ws_log.cell(row=row_idx, column=7, value=status_val)
        c7.alignment = Alignment(horizontal="center", vertical="center")
        c7.border = thin_border
        if status_val == "Passed":
            c7.font = pass_font
            c7.fill = pass_fill
        elif status_val == "Review":
            c7.font = review_font
            c7.fill = review_fill
        else:
            c7.font = flag_font
            c7.fill = flag_fill

        c8 = ws_log.cell(row=row_idx, column=8, value=row_data["Financial Issues"])
        c8.font = cell_font
        c8.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c8.border = thin_border

        c9 = ws_log.cell(row=row_idx, column=9, value=row_data["Compliance Issues"])
        c9.font = cell_font
        c9.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c9.border = thin_border

        risk_val = row_data["Vendor Risk Level"]
        c10 = ws_log.cell(row=row_idx, column=10, value=risk_val)
        c10.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c10.border = thin_border
        if risk_val.startswith("High Risk"):
            c10.font = risk_high_font
        elif risk_val.startswith("Moderate Risk"):
            c10.font = risk_moderate_font
        elif risk_val.startswith("Low Risk"):
            c10.font = risk_low_font
        else:
            c10.font = risk_default_font

        c11 = ws_log.cell(row=row_idx, column=11, value=row_data["Snippet Preview"])
        c11.font = Font(name="Calibri", size=9, italic=True, color="444444")
        c11.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c11.border = thin_border

    col_widths_log = {'A': 24, 'B': 22, 'C': 14, 'D': 14, 'E': 16, 'F': 16, 'G': 12, 'H': 26, 'I': 20, 'J': 24, 'K': 34}
    for col_letter, width in col_widths_log.items():
        ws_log.column_dimensions[col_letter].width = width

    ws_log.freeze_panes = "A5"

    saved_path = save_workbook_safely(wb, excel_output)

    print(f"========================================")
    print(f" Audit Complete! Multi-tab dashboard saved:")
    print(f" -> {os.path.abspath(saved_path)}")
    print(f"========================================")


if __name__ == "__main__":
    main()
